from openpyxl import load_workbook
from xlsx_to_json import *
import os.path

TEST_FILE= "sample/ResourceAssessmentSummaryData032011.xlsx"
EXPECTED_SHEET_NAMES = ['All Summary Data'] 

def test_showsheets():
    """
    The list of sheets of test file:
    """
    workbook = load_workbook(TEST_FILE, read_only=True)
    assert EXPECTED_SHEET_NAMES == workbook.sheetnames


def test_processsheet():
    workbook = load_workbook(TEST_FILE, read_only=True)
    sheet = EXPECTED_SHEET_NAMES[0]
    data = processsheet(workbook, sheet)
    assert len(data) == 191
    s = data[0]
    assert len(s) == 15

