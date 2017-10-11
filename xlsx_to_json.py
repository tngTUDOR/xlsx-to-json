"""XLSX -> json converter

The first row of the xlsx file has the attributes.

if no output file is provided, the name of the workbook (xlsx file) is used
and the name of the sheet is concatenated.
"""
import json
import os.path

import begin # rapid implementation of command line + main function
import tqdm # progress bars
from openpyxl import load_workbook

def processsheet(workbook, sheet):
    """ Transform sheet cells to json """
    worksheet = workbook[sheet]

    data = []
    keys = []
    print("Calculating page dimensions")
    dims = worksheet.calculate_dimension(force=True)
    print("Dimensions: "+dims)
    # Create the keys from the first row
    for row in worksheet.iter_rows(min_row=1, max_col=worksheet.max_column, max_row=1):
        for cell in row:
            keys.append(cell.value)

    for row_number, row in enumerate(tqdm.tqdm(worksheet.iter_rows(), desc="Row")):
        if row_number == 0:
            continue
        row_data = {}
        for col_number, cell in enumerate(tqdm.tqdm(row, desc="Cell")):
            row_data[keys[col_number]] = cell.value
        data.append(row_data)

    return data


def dumpdata(data, outputfile, indent):
    """ Dump the data into json format """
    with open(outputfile, 'w') as json_file:
        if indent:
            json_file.write(json.dumps({'data': data}, indent=4))
        else:
            json_file.write(json.dumps({'data': data}))

def processworkbook(inputfile, sheets, outputfile, indent):
    """ dump each specified sheet of the inputfile """
    workbook = load_workbook(inputfile, read_only=True)
    for sheet in sheets:
        if outputfile == '':
            outputfile = "".join([os.path.splitext(os.path.basename(inputfile))[0], sheet, '.json'])
        jsondata = processsheet(workbook, sheet)
        dumpdata(jsondata, outputfile, indent)

def showsheets(inputfile):
    """ Print the names of the sheets in the workbook"""
    workbook = load_workbook(inputfile, read_only=True)
    print("These are the sheets in the wb:")
    print(workbook.sheetnames)

@begin.start
def main(inputfile, outputfile='', indent=True, *sheet):
    """ run """
    assert (os.path.isfile(inputfile)), "Provided input file not found."
    if not sheet:
        showsheets(inputfile)
    else:
        processworkbook(inputfile, sheet, outputfile, indent)
